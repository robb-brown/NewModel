import numpy
import rpy2.robjects as robjects
import rpy2.robjects.vectors
import rpy2.rinterface as rinterface
from rpy2.robjects.packages import importr
import copy
import traceback
import collections
import math
from math import *
from bisect import bisect_left
import traceback
import datetime, time

import os.path as p

try:
	import openpyxl as xls
except:
	pass



def formatFloat(f,totalLength):
	if f == None or f == 'NA':
		return '-'.center(totalLength,' ')
	try:
		f = float(f)
	except:
		return '-'.center(totalLength,' ')
	if isnan(f):
		return '%*s' % (totalLength,'NaN')
	if f == 0.0:
		return '%*s' % (totalLength,'0.0')
	
	if f < 0:
		totalLength -= 1
	
	whole = len('%d' % int(round(f)))
	if -totalLength+1 < trunc(log10(abs(f))) <= totalLength-1:
		decimals = totalLength-whole-1
		if decimals < 1:
			s = '%*d' % (totalLength,round(f))
		else:
			s = ('%*.*f' % (totalLength,decimals,f)).rstrip('0')
			s = '%*s' % (totalLength,s)
	
	else:		# need scientific notation
		exponent = max([2,trunc(log10(abs(log10(abs(f))))) + 1])
		decimals = int(totalLength-2-2-exponent)
		if decimals < 1:
			s = '%*.0e' % (totalLength,f)
		else:
			s = '%*.*e' % (totalLength,decimals,f)
		
	return s


def sortkeypicker(keynames):
	negate = set()
	for i, k in enumerate(keynames):
		if k[:1] == '-':
			keynames[i] = k[1:]
			negate.add(k[1:])

	def getit(adict):
		composite = [adict[k] for k in keynames]
		for i, (k, v) in enumerate(zip(keynames, composite)):
			if k in negate:
				composite[i] = -v
		return composite
	return getit


def accumulateKeys(accumulator,sample,keys):
	for k in keys:
		if not k in accumulator:
			accumulator[k] = []
		try:
			accumulator[k].append(sample[k])
		except:
			pass
	return accumulator


	
def sortPK(data,primaryKey):
	if not primaryKey:
		return data
	data2 = {}
	for d in data:
		pkv = d[primaryKey]
		d.pop(primaryKey)
		data2[pkv] = d
	return data2



class ModelData(object):
	
	def __init__(self,data=None,keyName='key',style='stats',extras=None,order=None):
		self.vectorDict = None
		self.samples = None
		self.RData = None
		self.keySubstitutions = {}
		self.R = robjects.r
		self.defaults = None
		self.derivedValues = []
		self.references = {}
		self.order = order
		if data:
			self.setData(data,keyName,style,extras)
		
		
	def __str__(self,columns = None,floatWidth=5,columnSpacing=2):
		rdata = self.getRData()
		order = columns if columns else self.order
		columns = sorted(list(rdata.colnames))
		if order:
			columns = list([i.replace(' ','.').replace('-','.') for i in order]) + sorted(list(set(columns).difference(order)))
		else:
			columns = columns
				
		if self.order:
			cnames = list(rdata.colnames)
			sortKey = zip(*[list(rdata[cnames.index(i.replace(' ','.').replace('-','.'))]) for i in self.order])
			sortKey = numpy.array(sorted(range(len(sortKey)),key=sortKey.__getitem__))
		else:
			sortKey = None
		
		columnIndexes = [list(rdata.colnames).index(column) for column in columns]
		dataWidths = []; columnWidths = []; types = []
		for ci,column in enumerate(columns):
			c = columnIndexes[ci]
			if 'levels' in dir(rdata[c]):			# Factor vector
				widths = [len(i) for i in rdata[c].levels]
				try:
					dataWidths.append(max(widths))
				except:
					dataWidths.append(0)
				types.append('factor')
			else:									# Float vector
				dataWidths.append(floatWidth)
				types.append('float')

			columnWidths.append(int(max((dataWidths[-1],len(column)))))
			
		
		if sortKey is not None:
			data = [[formatFloat(f,columnWidths[c]) for f in numpy.array(rdata[rc])[sortKey]] if types[c] == 'float' else numpy.array(list(self.R['as.character'](rdata[rc])))[sortKey] for c,rc in enumerate(columnIndexes)]	
		else:
			data = [[formatFloat(f,columnWidths[c]) for f in rdata[rc]] if types[c] == 'float' else list(self.R['as.character'](rdata[rc])) for c,rc in enumerate(columnIndexes)]	
		s = (('%*s' % (columnSpacing,''))).join(['%*s' % (columnWidths[c],column.rjust(columnWidths[c])) for c,column in enumerate(columns)]) + '\n'
		s += '\n'.join([('%*s' % (columnSpacing,'')).join(['%*s' % (columnWidths[c],data[c][r]) for c,column in enumerate(columns)]) for r in range(0,len(rdata.rownames))])
		return s
	
		
	def addData(self,dictionary,keyName='key',style='stats',extras=None):
		if style == 'vector':
			# Don't remember ;)
			self.addVectorDict(dictionary,keyName)
		elif style in ['stats','samples']:
			# a list of dictionaries holding single samples
			self.addSamples(dictionary,extras=extras)
		elif style == 'list':
			# A nonstandard timecourse style list of dictionaries holding multiple samples
			self.addList(dictionary,extras)
		elif style == 'timecourse':
			# A standard lingo timecourse
			if dictionary.__class__ == {}.__class__:
				self.addTimecourseDict(dictionary,extras)
			else:
				self.addTimecourseList(dictionary,extras)
		else:
			print("Invalid data style")
		
		
	def setData(self,dictionary,keyName='key',style='stats',extras=None):
		self.samples = None
		self.vectorDict = None
		self.RData = None
		self.addData(dictionary,keyName,style,extras)
		
		
	def mergeData(self,newData,primaryKeys=None,createNew=False):
		samples1 = copy.deepcopy(self.samples)
		if newData.__class__ == self.__class__:
			samples2 = newData.samples
		else:
			samples2 = newData
		if primaryKeys == None:
			primaryKeys = set(samples1[0].keys()).intersection(set(samples2[0].keys()))
		keys1 = samples1[0].keys(); keys2 = samples2[0].keys()
		null1 = dict([(i,None) for i in samples1[0].keys() if not i in keys2])
		null2 = dict([(i,None) for i in samples2[0].keys() if not i in keys1])
		for i in samples1:
			i.update(null2) 
		pkstr = '-'.join(['%%(%s)s' % i for i in primaryKeys])
		samples1 = sorted(samples1,key=lambda elem: pkstr % elem); pk1 = [pkstr % i for i in samples1]
		samples2 = sorted(samples2,key=lambda elem: pkstr % elem); pk2 = [pkstr % i for i in samples2]
		for i,pk in enumerate(pk2):
			lo = bisect_left(pk1,pk)
			if lo < len(pk1) and pk1[lo] == pk:
				l = lo
				while l < len(pk1) and pk1[l] == pk:
					samples1[l].update(samples2[i])
					l+=1
			else:				# Our sample isn't in the first set so add it
				s = copy.deepcopy(samples2[i])
				s.update(null1)
				samples1.append(s)
		if createNew:
			return ModelData(samples1)
		else:
			self.setData(samples1)
			return self
		
		
	# *** Import/Export and load/save *** 	

	def writeXLSX(self,fname):
		data = self.samples

		columns = sorted(data[0].keys())
		if self.order:
			columns = list([i.replace(' ','.').replace('-','.') for i in order]) + sorted(list(set(columns).difference(order)))
		else:
			columns = columns
		
		headers = columns
		wb = xls.Workbook()
		ws = wb.get_active_sheet()
		for c,col in enumerate(headers):
			ws.cell(row=1,column=c+1,value = col)
			for r,row in enumerate(data):
				try:
					_ = ws.cell(row=r+2,column=c+1,value = row[col])
				except:
					print('\n',r, row.keys())
					traceback.print_exc()
		
		wb.save(fname)
		

	@classmethod
	def readXLSX(self,fname,sheet=0):
		wb = xls.load_workbook(fname,data_only=True)
		sheet = wb.worksheets[sheet]
		headers = []
		rows = list(sheet.rows)
		for header in rows[0]:
			headers.append(header.value)

		data = []
		for r,row in enumerate(rows[1:]):
			d = {}
			for c,col in enumerate(row):
				d[headers[c]] = col.value
	
			data.append(d)
		
		d = ModelData(data)
		d.order = headers
		return d
		
		
	@classmethod
	def read(self,fname,sheet=0):
		if p.splitext(fname)[-1] == '.xlsx':
			return self.readXLSX(fname,sheet)
		
	

	# *** END Import/Export and load/save *** 
			
			
	def setOrder(self,order):
		self.order = order
			
			
	def descriptive(self,key,weights=None):
		vector = numpy.array([i[key] for i in self.samples])
		try:
			if weights:
				weights = numpy.array([i[weights] for i in self.samples])
				weightedMean = float(sum(vector*weights)) / sum(weights)
			else:
				weightedMean = None
			return {'mean':numpy.average(vector),
					'std':numpy.std(vector),
					'weightedMean':weightedMean,
					'min':numpy.min(vector),
					'max':numpy.max(vector),
					'median':numpy.median(vector),
					'qtl1' : numpy.percentile(vector,25),
					'qtl3' : numpy.percentile(vector,75),
					'IQR' : numpy.percentile(vector,75) - numpy.percentile(vector,25),
					}
		except:
			return {k : len(numpy.where(vector == k)[0]) for k in numpy.unique(vector)}
		
				
	def setReference(self,factor,reference):
		self.references[factor] = reference
		self.convertToRData()
		
		
	def setReferences(self,references):
		self.references.update(references)
		self.convertToRData()
		
		
	def clearReferences(self):
		self.references = {}
		self.convertToRData()
		
			
	def addVectorDict(self,d,keyName='key'):
		self.vectorDict = dict(d)
		self.convertToSamples(keyName)

		
	def addSamples(self,samples,extras=None):
		if extras:
			for s in samples:
				for e in extras.keys():
					s[e] = extras[e]
		
		if self.samples == None:
			self.samples = []
		self.samples += list(samples)
		self.convertToRData()
		
		
	def addList(self,l,extras=None):
		self.convertFromList(l,extras=extras)
		
		
	def addTimecourseList(self,d,extras=None):
		self.convertFromTimecourseList(list(d),extras)
	
	def addTimecourseDict(self,d,extras=None):
		self.convertFromTimecourseDict(dict(d),extras)
				
		
	def getVectorData(self):
		return dict(self.vectorDict)
		
		
	def getStatData(self):
		return copy.deepcopy(self.samples)
		
		
	def getRData(self,mustHave=[]):
#		if order:
#			self.order = order
#		if self.order:
#			print "Getting RData.  Order is: %s" % self.order
#			s = ','.join(self.order)
#			robjects.globalenv["rdata"] = robjects.DataFrame(copy.deepcopy(self.RData))
#			rdata = self.R('rdata[with(rdata,order(%s)),]' % s)
#			self.R['rm']("rdata")
#			return rdata
#		else:

		if getattr(self,'RData',None) is None or self.RData is None:
			self.convertToRData()
		rdata = robjects.DataFrame(copy.deepcopy(self.RData))
		
		if mustHave:
#			import code; code.interact(local=locals())
			robjects.globalenv["rdataData"] = rdata
			effectsString = ','.join(['"%s"' % i for i in mustHave])
			rdata = self.R('rdataData[complete.cases(rdataData[,c(%s)]),]' % effectsString)
		return rdata
		
		
	def setRData(self,rdata):
		self.RData = rdata
		
		
	def convertToSamples(self,keyName='key'):
		self.samples = []
		d = self.vectorDict
		samples = []
		for k in d.keys():
			d[k][keyName] = numpy.repeat([k],len(d[k].values()[0]))
			s = zip(*d[k].values())
			s = [dict(zip(d[k].keys(),i)) for i in s]
			samples.extend(s)
#		self.samples = dict([(i,samples[i]) for i in range(0,len(samples))])
		self.samples = samples
		self.convertToRData()
		
				
	def convertToVectorDict(self):
		pass
		
		
	def makeEstimateTimecourse(self,contrast,elapsed,subject='',mask='',lesion='',timepoint='',timepointElapsed=0,volume=0,xKey='elapsed'):
		tc = {'elapsed':elapsed, 'contrast':contrast, 'subject':subject, 'mask':mask,'lesion':lesion,'timepoint':timepoint,'timepointElapsed':timepointElapsed,'volume':volume}
		tc[xKey] = elapsed
		tc['timecourse'] = numpy.zeros(len(elapsed))
		self.convertFromTimecourseList([tc])
		
		
	def makeEstimate(self,yKey,xKey,x,individual=None,extraKeys={}):
		tc = dict([(k,None) for k in extraKeys])
		tc.update({xKey : x, yKey : numpy.zeros(len(x))})
		if individual:
			if individual.__class__ == str:
				individual = [individual,'subject1']
			individual = list(individual)
			if individual[1].__class__ == str:
				individual[1] = numpy.repeat(individual[1],len(x))
			tc.update({individual[0]:individual[1]})
		self.convertFromList([tc])
		
		
	def convertFromList(self,data,extras=None):
		if extras == None:
			extras = {}
		samples = []
		
		for t,tc in enumerate(data):
			# Scan for sequences.  All sequences should be the same length
			slength = None
			fail = False
			for k in tc.keys():
				if (isinstance(tc[k],collections.Iterable)) and not isinstance(tc[k],str):
					if not slength:
						slength = len(tc[k])
					else:
						if not len(tc[k]) == slength:
							print("Item %d contains lists of different lengths!  Failing!" % t)
							fail = True
							break
			
			if not fail:
				for s in range(0,slength):
					sample = {}
					for k in tc.keys():
						if (not isinstance(tc[k],collections.Iterable)) or isinstance(tc[k],str):
							sample[k] = tc[k]
						else:
							sample[k] = tc[k][s]
		
					samples.append(sample)
		self.addSamples(samples)
					
		
	
	def convertFromTimecourseList(self,data,extras=None):
		# converts from a standard timecourse list to a sample dictionary
		sdata = []
		scount = 0
		if extras == None:
			extras = {}
			
		for t,tc in enumerate(data):
			if 'timepoints' in tc:
				timepoints = tc['timepoints']
			else:
				timepoints = numpy.array([None]).repeat(len(tc['elapsed']))
			for tp,e in enumerate(tc['elapsed']):
				elapsed = tc['elapsed'][tp]
				value = tc['timecourse'][tp]
				try:
					timepoint = timepoints[tp]
				except:
					print("Error - should be timepoints, but none found, timecourse %d." % t)
					traceback.print_exc()
					print("\n")
					timepoint = None
				if not elapsed is numpy.ma.masked and not value is numpy.ma.masked:
					sample = {}
					for k in tc.keys():
						if (not isinstance(tc[k],collections.Iterable)) or isinstance(tc[k],str):
							sample[k] = tc[k]
						else:
							sample[k] = tc[k][tp]
							
					sample['subject'] = tc['subject']
					sample['tissue'] = tc['mask']
					sample['lesionID'] = sample['subject']+'_l'+tc['lesion']
					sample['timepoint'] = tc['timepoint']
					sample['timepointElapsed'] = tc['timepointElapsed']
					sample['contrast'] = tc['contrast']
					sample['volume'] = tc['volume']
#					sample['ltime'] = elapsed
					sample['elapsed'] = elapsed
					sample['value'] = value
					sample[tc['contrast']] = value
					if timepoint:
						sample['dataTimepoint'] = timepoint
					for extra in extras.keys():
						sample[extra] = extras[extra]
					sdata.append(sample)
					scount += 1
		self.addSamples(sdata)
		  
			
	
	def convertFromTimecourseDict(self,data,extras=None):
		# converts from a standard timecourse dictionary to a sample dictionary
		sdata = []
		scount = 0
		if extras == None:
			extras = {}
		for s,subject in enumerate(data['subjects']):			
			for t,time in enumerate(data['times'][s]):
				if not data['timecourses'][s].mask[t]:
					sample = {}
					sample['subject'] = subject
					lid = data['lesionids'][s]
					try:
						lesionid = int(lid.split('_')[-1])
						lesionType = '_'.join(lid.split('_')[0:-1])
					except:
						lesionid = None
						lesionType = lid
					sample['tissue'] = lesionType
					sample['lesionID'] = subject+'l'+str(lesionid)
					sample['ltime'] = time
					sample['volume'] = data['volumes'][s]
					sample['MTR'] = data['timecourses'][s][t]
					sample['timepoint'] = data['timepoints'][s][t]
					if (t < len(data['meanTimes'])):
						sample['meanLTime'] = data['meanTimes'][t]
					else:
						sample['meanLTime'] = None
					for extra in extras.keys():
						sample[extra] = extras[extra]
					sdata.append(sample)
					scount += 1
		self.addSamples(sdata)
		
		
	def calculateDerivedValues(self,instructions,variables=None,verbose=False):
		for sample in range(0,len(self.samples)):
			for instruction in instructions:
				sval = dict([(k.replace('-','_'),v) for k,v in self.samples[sample].items()])
				locs = sval
				if variables: locs.update(variables)
				if 'variables' in instruction: locs.update(instruction.get('variables',{}))
				try:
					result = eval(instruction['criteria'],locs,globals())
				except:
					if verbose:
						print("Error evaluating derived values instruction %s" % instruction['criteria'])
						print("with data %s" % (sval))
						print()
						traceback.print_exc()
					result = None
				rtest = False
				try:
					if not result.__class__ == True.__class__:
						rtest = float(result)
						rtest = True
				except:
					rtest = False
				try:
					if result in instruction:			# We have a key for the result
						if not instruction[result] == None:			# If the result is not None, just set it
							self.samples[sample][instruction['key']] = instruction[result]
						else:				# If it's None, we only overwrite the existing key if it is missing or None
							if not self.samples[sample].get(instruction['key'],None):
								self.samples[sample][instruction['key']] = instruction[result]
					elif result or rtest:				# WE don't have an instruction key.  If the result is something, put it in directly
						self.samples[sample][instruction['key']] = result
					elif not instruction['key'] in self.samples[sample]:			# Otherwise, set it to none, if the existing one is missing
						self.samples[sample][instruction['key']] = None
				except:
					if verbose:
						print("Error evaluating result of derived values instruction %s" % instruction['criteria'])
						print("with data %s" % (self.samples[sample]))
						print()
						traceback.print_exc()
						
		self.convertToRData()
		toAdd = []
		for i in instructions:			
			try:
				if not i.get('doNotStore',False) and not i in self.derivedValues:
					toAdd.append(i)
			except:
				if not i.get('doNotStore',False) and not i['key'] in [x['key'] for x in self.derivedValues]:
					toAdd.append(i)
		self.derivedValues.extend(toAdd)
		return self
		
		
	def dropKeys(self,keys):
		for s in self.samples:
			for k in keys:
				try:
					s.pop(k)
				except:
					pass
		for k in keys:
			if k in self.references:
				self.references.pop(k)
		self.convertToRData()
		return self
		
		
	def filter(self,filter,vars=None):
		loc = {'samples':self.samples,'numpy':numpy,'math':math}
		if vars: loc.update(vars)
		self.samples = eval(filter,loc)
		self.convertToRData()
		
	def filteredCopy(self,filter,vars=None):
		temp = copy.deepcopy(self)
		temp.filter(filter,vars)
		return temp
		
		
	def filtered(self,filter,vars=None):
		return self.filteredCopy(filter,vars)
		
		
	def setDefaults(self,defaults):
		self.defaults = defaults
		for k,v in defaults.items():
			self.RData = self.R.within(self.RData,'%s <- relevel(%s,ref="%s")' % (k,k,v))
			
	@classmethod
	def convertColumnsToRData(self,data):
		try:
			for k in list(data.keys()):
				try:
					types = list(set(numpy.array([i.__class__ for i in data[k] if not i.__class__ == None.__class__])))
					if len(types) == 0:
						type = str
					elif len(types) == 1:
						type = types[0]
					elif float in types or numpy.float32 in types:
							type = float
					elif numpy.float64 in types:
							type = numpy.float64
					elif len(types) == 2 and str in types and u''.__class__ in types:
						type = u''.__class__
					else:
						print("There are multiple types in column %s (%s).  Converting everything to text.\n" % (k,types))
						type = None

					if data[k].__class__ == numpy.ma.MaskedArray:
						data[k] = data[k].filled(numpy.nan)
					else:
						data[k] = numpy.array(data[k])

					if type == str:
						data[k][numpy.equal(data[k],None)] = robjects.NA_Character
						data[k] = robjects.StrVector(data[k])						
					elif type == u''.__class__:
						data[k][numpy.equal(data[k],None)] == robjects.NA_Character
						data[k] = robjects.StrVector([i.encode('ascii','ignore') if i is not None else None for i in data[k]])
					elif type in [''.__class__,u''.__class__,numpy.array(['',''])[0].__class__,numpy.array([u'',u''])[0].__class__]:
						data[k][numpy.equal(data[k],None)] = robjects.NA_Character
						data[k] = robjects.StrVector(data[k])
					elif type in [int,numpy.int8,numpy.uint8,numpy.int16,numpy.uint16,numpy.int32,numpy.uint32,numpy.int64,numpy.uint64]:
						data[k][numpy.equal(data[k],None)] = robjects.NA_Integer
						data[k] = robjects.IntVector(data[k])
					elif type in [numpy.bool_,numpy.bool]:
						data[k] = data[k].astype(numpy.uint8)
						data[k][numpy.equal(data[k],None)] = robjects.NA_Integer
						data[k] = robjects.IntVector(data[k])
					elif type in [datetime.datetime]:
						data[k] = robjects.vectors.POSIXct(robjects.FloatVector([time.mktime(i.timetuple()) if i is not None else robjects.NA_Real for i in data[k]]))
					elif type == None:
						data[k] = robjects.StrVector([str(s) for s in data[k]])
					else:
						data[k] = data[k].astype(numpy.float64)
						data[k][numpy.equal(data[k],None)] = robjects.NA_Real
						data[k] = robjects.FloatVector(data[k])
				except:
					print("Exception on %s.  Skipping column %s" % (k,k))
					traceback.print_exc()
					data.pop(k)
					print()
		except:
			traceback.print_exc()
			print(k)
			print(data[k])
			print()
		return data
		
	
	def convertToRData(self,keySubstitutions=None):
		if keySubstitutions:
			self.keySubstutitions = keySubstitutions
		# make sure all the samples have the same keys
		data = {}
		allKeys = [s.keys() for s in self.samples]
		allKeys = set([item for sublist in allKeys for item in sublist])
		for k in allKeys:
			if k in self.keySubstitutions.keys():
				newK = str(self.keySubstitutions[k])
			else:
				newK = str(k)
			data[newK] = [i.get(k,None) for i in self.samples]
		self.convertColumnsToRData(data)
		
		# new Rpy2
		for col in data.keys():
			if isinstance(data[col],rpy2.robjects.vectors.StrVector):
				data[col] = self.R.factor(data[col])
		
		self.RData = self.R['data.frame'](**data)		
		for factor in self.references:
			try:
				i = self.RData.colnames.index(factor)
				if self.references[factor] in self.RData[i].levels:
					newlevels = list(self.RData[i].levels)
					newlevels.remove(self.references[factor])
					newlevels.append(self.references[factor])
					self.RData[i] = self.R.factor(self.RData[i],levels=newlevels)
				else:
					print("%s is not a valid level for %s" % (self.references[factor],factor))
			except:
				print("Releveling could not find factor %s" % factor)
				traceback.print_exc()
		
		
	def convertRDataToSamples(self):
		rdata = self.getRData()
		headers = list(rdata.names)
		samples = []
		for r in range(0,numpy.shape(rdata)[1]):
			sample = {}
			for c,col in enumerate(headers):
				if self.R['class'](rdata[c])[0] == 'factor':
					sample[col] = list(rdata[c].levels)[rdata[c][r]-1]
				else:
					sample[col] = rdata[c][r]
			samples.append(sample)
		self.samples = samples
		
		
	def getSampleIndex(self,sample,primaryKeys=None):
		primaryKeys = self.samples[0].keys() if primaryKeys is None else primaryKeys
		for s,samp in enumerate(self.samples):
			match = True
			for key in primaryKeys:
				if not samp[key] == sample[key]:
					match = False
					break
			if match:
				break
		return s if match else None
				
		
	def averageOver(self,overKey,avgKey=[],sumKey=[],differenceKey=[],ignoreKey=[],weightKey=None,commit=True,reverseOrder=False):
		# Average samples.  The overKey is the one to average over.  avgKeys will be averaged, sumKeys summed,
		#  ignoreKeys ignored. The average is weighted by the weightKey, if any.  Commit means to actually
		# put the result back into the data struture, otherwise it's returned as samples.
		# For example, to calculate the drop we want to average the MTR/value over elapsed, possibly ignoring things
		# like the lesionID.  Then we can subtract the postLesion=0 values from the postLesion=1s.
		
		if overKey.__class__ == str:
			overKey = [overKey]
		samples = copy.deepcopy(self.samples)
		constant = set(samples[0].keys()).difference(set(avgKey+sumKey+differenceKey+overKey+ignoreKey))
		condition = ''
		key = '('
		for k in constant:
			condition += "and (exemplar['%s'] == sample['%s']) " % (k,k)
			key += "itemgetter('%s')," % k

		key += ')'
		condition = condition[4:]
		samples = sorted(samples,key=sortkeypicker(constant),reverse=reverseOrder)
		samples.append(copy.deepcopy(samples[0])); samples[-1][overKey[0]] = None				# Null sample at the end to clear the accumulator

		newSamples = []
		exemplar = dict(samples[0])
		weights = []
		accumulator = accumulateKeys({},exemplar,avgKey+sumKey+differenceKey)
		if weightKey:
			weights.append(exemplar[weightKey])
		for sample in samples[1:]:
			if eval(condition):
				accumulator = accumulateKeys(accumulator,sample,avgKey+sumKey+differenceKey)
				if weightKey:
					weights.append(sample[weightKey])
			else:
				for k in avgKey:
					if weightKey:
						exemplar[k] = numpy.average(accumulator[k],weights=weights)
					else:
						exemplar[k] = numpy.average(accumulator[k])
				for k in sumKey:
					exemplar[k] = numpy.sum(accumulator[k])
				for k in differenceKey:
					if len(accumulator[k]) == 2:
						exemplar[k] = accumulator[k][1] - accumulator[k][0]
					elif len(accumulator) < 2:
						print("Less than two values for difference key.")
						exemplar[k] = None
					else:
						print("More than two values for difference key!")
						exemplar[k] = None
				for k in overKey:
					temp = exemplar.pop(k)
				newSamples.append(exemplar)
				exemplar = sample
				accumulator = accumulateKeys({},exemplar,avgKey+sumKey+differenceKey)
				if weightKey:
					weights = [sample[weightKey]]
		if commit:
			self.samples = newSamples
			self.convertToRData()
		else:
			return newSamples

			

	def accumulate(self,overKey,accKey=[],ignoreKey=[],constant=[],returnSampleIndex=False):
		samples = copy.deepcopy(self.samples)
		if constant == []:
			constant = list(set(samples[0].keys()).difference(set(accKey+overKey+ignoreKey)))
		elif ignoreKey == []:
			ignoreKey = list(set(samples[0].keys()).difference(set(accKey+overKey+constant)))
		
		if len(constant) == 0:
			samples = [dict(s.items()+[('constant',1)]) for s in samples]
			constant.append('constant')
			
		condition = ''
		key = '('
		for k in constant:
			condition += "and (exemplar['%s'] == sample['%s']) " % (k,k)
			key += "itemgetter('%s')," % k

		key += ')'
		condition = condition[4:]
		samples = sorted(samples,key=sortkeypicker(constant+overKey))
		try:
			samples.append(copy.deepcopy(samples[0])); 
		except:
			# This is probably a numpy bug where a masked element seems to get treated as an array sometimes
			samples.append(dict([(k,copy.deepcopy(v) if not v.__class__ == numpy.ma.masked.__class__ else numpy.ma.masked) for k,v in samples[0].items()]))
			
			
		samples[-1][constant[0]] = None				# Null sample at the end to clear the accumulator

		newSamples = []
		exemplar = dict(samples[0])
		accumulator = accumulateKeys({},exemplar,accKey)
		sampleIndex = [0]
		for sample in samples[1:]:
			if eval(condition):
				accumulator = accumulateKeys(accumulator,sample,accKey)
				sampleIndex.append(len(newSamples))
			else:
				for k in overKey:
					temp = exemplar.pop(k)
				for k in ignoreKey:
					temp = exemplar.pop(k)
				for k in accKey:
					exemplar[k] = accumulator[k]
				newSamples.append(exemplar)
				exemplar = sample
				sampleIndex.append(len(newSamples))
				accumulator = accumulateKeys({},exemplar,accKey)
		if returnSampleIndex:
			return newSamples,sampleIndex
		else:
			return newSamples
